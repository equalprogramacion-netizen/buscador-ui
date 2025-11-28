# ============================== 
# IMPORTACIÓN DE LIBRERÍAS
# ==============================
from flask import Flask, render_template, request, send_file, session, jsonify, redirect, make_response, g, Response
import mysql.connector
from mysql.connector import pooling
from pyproj import Transformer
import csv, io, os, glob, time, uuid, hmac, json, base64
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from openpyxl.utils import get_column_letter
from typing import cast
from dotenv import load_dotenv

load_dotenv()  # Cargar variables del .env

# ==============================
# CONFIGURACIÓN GENERAL DE FLASK
# ==============================
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "cambia-esta-clave")  # solo para session de Flask (export paths)
app.config['EXPORT_FOLDER'] = os.getenv("EXPORT_FOLDER", 'temp_exports')
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)

IS_PROD = os.getenv("ENV", "development").lower() == "production"

# ==============================
# BRANDING (colores y estilos)
# ==============================
def brand_vars():
    return {
        'PRIMARY': os.getenv('PRIMARY', '#22c55e'),
        'ACCENT': os.getenv('ACCENT', '#7c9bff'),
        'INK': os.getenv('INK', '#e6ebff'),
        'MUTED': os.getenv('MUTED', '#9aa3c7'),
        'BG': os.getenv('BG', '#0b1020'),
        'CARD': os.getenv('CARD', '#111833'),
        'BORDER': os.getenv('BORDER', '#1f2547'),
        'LOGO_RADIUS': os.getenv('LOGO_RADIUS', '16px')
    }
# ==========================================================
# PUERTA (SSO LIGERO) DESDE EL HUB
# ==========================================================
GATEWAY_SHARED_SECRET      = os.getenv("GATEWAY_SHARED_SECRET", "cambia-esto-por-un-secreto-largo-y-unico")
GATEWAY_SHARED_SECRET_PREV = os.getenv("GATEWAY_SHARED_SECRET_PREV", "")  # rotación opcional
GATE_AUD = os.getenv("GATE_AUD", "buscador")  # audiencia esperada en este servicio

# Cookie de sesión local de este servicio (no es la del Hub)
SVC_SESSION_COOKIE = os.getenv("SVC_SESSION_COOKIE", "svc_buscador")
SVC_SESSION_TTL    = int(os.getenv("SVC_SESSION_TTL", "1800"))  # 30 min

# Dónde enviar al usuario si no tiene sesión aquí
HUB_HOME = os.getenv("HUB_HOME", "http://127.0.0.1:8000/choose")

# Rutas anónimas permitidas (raíz protegida)
ANON_PATHS = set((
    # "/"  # ← quitado: la raíz ahora exige SSO vía Hub
    "/health", "/healthz",
    "/favicon.ico", "/robots.txt",
    # Si NO quieres exponer docs del buscador, no agregues aquí sus rutas
))

# ============ Helpers base64url/HMAC para `st` ============
def _b64url_pad(s: str) -> bytes:
    s += "=" * ((4 - len(s) % 4) % 4)
    return s.encode("ascii")

def _b64url_decode_to_json(b64: str) -> dict:
    raw = base64.urlsafe_b64decode(_b64url_pad(b64))
    return json.loads(raw.decode("utf-8"))

def _sign_st_payload(payload_b64: str, secret: str) -> str:
    sig = hmac.new(secret.encode("utf-8"), payload_b64.encode("ascii"), digestmod="sha256").digest()
    return base64.urlsafe_b64encode(sig).rstrip(b"=").decode("ascii")

def _verify_st(token: str) -> dict | None:
    """
    st = <base64url(json_payload)>.<base64url(signature)>
    payload = {"sub","aud","iat","exp","rid","iss"}  (como lo genera el Hub)
    Valida HMAC (secret actual o previo), exp y aud.
    Devuelve dict si es válido; si no, None.
    """
    if not token or "." not in token:
        return None
    parts = token.split(".")
    if len(parts) != 2:
        return None
    payload_b64, sig_b64 = parts[0], parts[1]

    good_sig = _sign_st_payload(payload_b64, GATEWAY_SHARED_SECRET)
    if sig_b64 != good_sig:
        if not GATEWAY_SHARED_SECRET_PREV:
            return None
        good_sig_prev = _sign_st_payload(payload_b64, GATEWAY_SHARED_SECRET_PREV)
        if sig_b64 != good_sig_prev:
            return None

    try:
        payload = _b64url_decode_to_json(payload_b64)
    except Exception:
        return None

    now = int(time.time())
    # exp/iat mínimos (exp requerido, iat opcional)
    try:
        if int(payload.get("exp", 0)) < now:
            return None
    except Exception:
        return None

    if payload.get("aud") != GATE_AUD:
        return None

    return payload

# ============ Cookie de sesión local firmada ============
# Podemos reutilizar el secret del gateway; si prefieres, define BUSCADOR_SESSION_SECRET aparte.
from itsdangerous import TimestampSigner, BadSignature, SignatureExpired
_svc_signer = TimestampSigner(os.getenv("BUSCADOR_SESSION_SECRET", GATEWAY_SHARED_SECRET))

def _set_svc_session(resp, email: str):
    token = _svc_signer.sign(email.encode("utf-8")).decode("utf-8")
    resp.set_cookie(
        SVC_SESSION_COOKIE, token,
        max_age=SVC_SESSION_TTL, httponly=True,
        samesite="strict" if IS_PROD else "lax",
        secure=IS_PROD, path="/"
    )

def _get_svc_email() -> str | None:
    tok = request.cookies.get(SVC_SESSION_COOKIE)
    if not tok:
        return None
    try:
        raw = _svc_signer.unsign(tok, max_age=SVC_SESSION_TTL)
        return raw.decode("utf-8")
    except (BadSignature, SignatureExpired):
        return None

def _clear_svc_session(resp):
    resp.delete_cookie(SVC_SESSION_COOKIE, path="/")

# ============ Guard (antes y después de cada request) ============
@app.before_request
def gate_guard():
    # En desarrollo, desactivar el guard para permitir pruebas locales
    if not IS_PROD:
        return
    path = (request.path or "").rstrip("/")

    # permite estáticos (Flask sirve /static/* automáticamente)
    if path.startswith("/static/"):
        return

    # permite rutas anónimas exactas
    if path in ANON_PATHS:
        return

    # ¿ya hay sesión local?
    email = _get_svc_email()
    if email:
        g._svc_email = email
        return

    # ¿viene st en query o Authorization Bearer?
    st = request.args.get("st")
    if not st:
        auth = request.headers.get("Authorization", "")
        if auth.lower().startswith("bearer "):
            st = auth.split(" ", 1)[1].strip()

    payload = _verify_st(st) if st else None
    if payload:
        # marcar para setear cookie al final
        g._set_cookie_email = payload.get("sub", "")
        return

    # No autorizado → 401 con botón al Hub
    html = f"""
    <!doctype html><html><head><meta charset="utf-8"/>
      <title>401 — Autenticación requerida</title>
    </head>
    <body style="font-family:system-ui;background:#0b1020;color:#e6ebff;display:grid;place-items:center;height:100vh;margin:0">
      <div style="max-width:680px;background:#0f162b;border:1px solid rgba(255,255,255,.08);padding:24px;border-radius:14px">
        <h2 style="margin:0 0 8px">Acceso restringido</h2>
        <p style="margin:0 0 14px;opacity:.8">Para usar el Buscador debes entrar desde el Hub.</p>
        <a href="{HUB_HOME}" style="display:inline-block;background:#22c55e;color:#08150c;padding:10px 16px;border-radius:10px;font-weight:800;text-decoration:none">Ir al Hub</a>
      </div>
    </body></html>
    """
    return make_response(html, 401)

@app.after_request
def gate_after(resp):
    # Si entró con st válido en esta request, crea cookie de sesión local
    email = getattr(g, "_set_cookie_email", None)
    if email:
        _set_svc_session(resp, email)

    # Evitar cache de HTML por defecto
    ct = (resp.headers.get("Content-Type") or "").lower()
    if "text/html" in ct:
        resp.headers["Cache-Control"] = "no-store"
    # Cache extendida para assets estáticos (siempre que pasen por Flask)
    if request.path.startswith("/static/"):
        resp.headers["Cache-Control"] = "public, max-age=31536000, immutable"
    return resp

# Logout local del buscador
@app.route("/logout")
def logout_local():
    resp = redirect(HUB_HOME)
    _clear_svc_session(resp)
    resp.headers["Clear-Site-Data"] = '"cache","cookies","storage"'
    return resp

# ==========================================================
# FUNCIÓN PARA LIMPIAR ARCHIVOS ANTIGUOS (CSV Y EXCEL > 1h)
# ==========================================================
def limpiar_archivos_antiguos():
    una_hora_atras = time.time() - 3600
    for extension in ('*.csv', '*.xlsx'):
        for archivo in glob.glob(os.path.join(app.config['EXPORT_FOLDER'], extension)):
            if os.path.getmtime(archivo) < una_hora_atras:
                try:
                    os.remove(archivo)
                except:
                    pass

# ========================================
# CONFIGURACIÓN DEL POOL DE CONEXIONES DB
# ========================================
# OJO: el schema (base de datos) va en DB_NAME; la tabla sola en DB_TABLE
DB_TABLE = os.getenv("DB_TABLE", "biotic_database").strip()   # nombre de la tabla
DB_NAME  = os.getenv("DB_NAME",  "railway").strip()           # schema/base de datos

# Validación mínima de ENV
_required = ["DB_HOST", "DB_PORT", "DB_USER", "DB_PASSWORD"]
missing = [k for k in _required if not os.getenv(k)]
if missing:
    raise RuntimeError(f"Faltan variables en .env: {', '.join(missing)}")

dbconfig = {
    "host": os.getenv("DB_HOST"),
    "port": int(os.getenv("DB_PORT", "3306")),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
    "database": DB_NAME,
    "charset": "utf8mb4",
    "use_unicode": True,
    "autocommit": True,
    "connection_timeout": 10,
}
cnxpool = pooling.MySQLConnectionPool(pool_name="mypool", pool_size=5, **dbconfig)

# Helper para calificar nombres con backticks
def tq(schema: str, table: str) -> str:
    return f"`{schema}`.`{table}`"

FULL_TABLE = tq(DB_NAME, DB_TABLE)

# ================================================
# CACHÉ Y FUNCIÓN PARA OBTENER NOMBRES DE COLUMNAS
# ================================================
columnas_cache = []

def obtener_columnas():
    """Lee y cachea las columnas de la tabla a consultar."""
    global columnas_cache
    if not columnas_cache:
        conn = cnxpool.get_connection()
        cursor = conn.cursor()
        cursor.execute(f"SHOW COLUMNS FROM {FULL_TABLE}")
        columnas_cache = [col[0] for col in cursor.fetchall()]
        cursor.close()
        conn.close()
    return columnas_cache

# ===============================
# RUTA PRINCIPAL "/"
# ===============================
@app.route('/')
def index():
    conn = cnxpool.get_connection()
    cursor = conn.cursor()

    cursor.execute(f"SELECT DISTINCT Municipio FROM {FULL_TABLE} ORDER BY Municipio")
    municipios = [row[0] for row in cursor.fetchall()]

    cursor.execute(f"SELECT DISTINCT Proyecto FROM {FULL_TABLE} ORDER BY Proyecto")
    proyectos = [row[0] for row in cursor.fetchall()]

    cursor.execute(f"SELECT DISTINCT Nombre_cientifico FROM {FULL_TABLE} ORDER BY Nombre_cientifico")
    especies = [row[0] for row in cursor.fetchall()]

    cursor.execute(
        f"SELECT DISTINCT Grupo_Biologico FROM {FULL_TABLE} "
        f"WHERE Grupo_Biologico IS NOT NULL ORDER BY Grupo_Biologico"
    )
    grupos_biologicos = [row[0] for row in cursor.fetchall()]

    cursor.execute(
        f"SELECT DISTINCT Tipo_Hidrobiota FROM {FULL_TABLE} "
        f"WHERE Tipo_Hidrobiota IS NOT NULL ORDER BY Tipo_Hidrobiota"
    )
    tipos_hidrobiota = [row[0] for row in cursor.fetchall()]

    cursor.close()
    conn.close()

    return render_template(
        'index.html',
        columnas=obtener_columnas(),
        municipios=municipios,
        proyectos=proyectos,
        especies=especies,
        grupos_biologicos=grupos_biologicos,
        tipos_hidrobiota=tipos_hidrobiota,
        **brand_vars()
    )

# ===============================
# RUTA "/columnas"
# ===============================
@app.route("/columnas")
def columnas():
    return jsonify({"columnas": obtener_columnas()})

# ======================================
# RUTA "/buscar" (BÚSQUEDA DE REGISTROS)
# ======================================
@app.route('/buscar', methods=['POST'])
def buscar():
    limpiar_archivos_antiguos()

    # ---------- Captura de filtros ----------
    palabra_clave = request.form.get('palabra', '').strip()
    columna_clave = request.form.get('columna', '')
    municipio = request.form.get('filtro_municipio')
    proyecto = request.form.get('filtro_proyecto')
    nombre_comun = request.form.get('filtro_nombre_comun')
    nombre_cientifico = request.form.get('filtro_especie')
    codigo_de_muestra = request.form.get('codigo_de_muestra')
    grupo_biologico = request.form.get('filtro_grupo_biologico')
    tipo_hidrobiota = request.form.get('filtro_tipo_hidrobiota')
    columnas_mostrar = request.form.getlist('columnas_mostrar')

    # ---------- Columnas ----------
    columnas_disponibles = obtener_columnas()

    # columnas clave que siempre incluimos en la exportación
    columnas_clave = [
        'Nombre_cientifico',
        'Nombre_comun',
        'Codigo_de_muestra',
        'Proyecto',
        'Fecha_de_colecta'
    ]

    # Si no se seleccionaron o pidieron __todas__
    if '__todas__' in columnas_mostrar or not columnas_mostrar:
        columnas_mostrar = columnas_disponibles.copy()
    else:
        columnas_mostrar = list(set(columnas_mostrar + columnas_clave))

    # Validamos contra columnas reales (respeta el orden de la tabla)
    columnas_mostrar = [col for col in columnas_disponibles if col in columnas_mostrar]

    # Columnas mínimas necesarias para procesamiento (mapa y popups)
    columnas_min_proc = {
        'Latitud_decimal', 'Longitud_decimal', 'Codigo_EPSG_decimal',
        'Nombre_cientifico', 'Nombre_comun', 'Codigo_de_muestra', 'Proyecto', 'Fecha_de_colecta'
    }
    columnas_requeridas = set(columnas_mostrar) | columnas_min_proc
    # Selección final respetando el orden real de la tabla
    columnas_select = [c for c in columnas_disponibles if c in columnas_requeridas]

    # Helper para citar columnas con backticks
    def qc(col: str) -> str:
        return f"`{col}`"

    # ---------- Construcción del SQL ----------
    conn = cnxpool.get_connection()
    cursor = conn.cursor(dictionary=True)

    # SELECT explícito para evitar confusiones de columnas
    select_cols_sql = ", ".join(qc(c) for c in columnas_select)
    query = f"SELECT {select_cols_sql} FROM {FULL_TABLE} WHERE 1=1"
    filtros = []
    valores = []

    if municipio:
        filtros.append("Municipio LIKE %s")
        valores.append(f"%{municipio}%")
    if proyecto:
        filtros.append("Proyecto LIKE %s")
        valores.append(f"%{proyecto}%")
    if nombre_comun:
        filtros.append("Nombre_comun LIKE %s")
        valores.append(f"%{nombre_comun}%")
    if nombre_cientifico:
        filtros.append("Nombre_cientifico LIKE %s")
        valores.append(f"%{nombre_cientifico}%")
    if codigo_de_muestra:
        filtros.append("Codigo_de_muestra LIKE %s")
        valores.append(f"%{codigo_de_muestra}%")
    if grupo_biologico:
        filtros.append("Grupo_Biologico LIKE %s")
        valores.append(f"%{grupo_biologico}%")
    if tipo_hidrobiota:
        filtros.append("Tipo_Hidrobiota LIKE %s")
        valores.append(f"%{tipo_hidrobiota}%")
    if palabra_clave:
        if columna_clave and columna_clave != "__todas__" and columna_clave in columnas_disponibles:
            filtros.append(f"{qc(columna_clave)} LIKE %s")
            valores.append(f"%{palabra_clave}%")
        elif columna_clave == "__todas__" and columnas_disponibles:
            subfiltros = [f"{qc(col)} LIKE %s" for col in columnas_disponibles]
            filtros.append("(" + " OR ".join(subfiltros) + ")")
            valores.extend([f"%{palabra_clave}%"] * len(columnas_disponibles))

    if filtros:
        query += " AND " + " AND ".join(filtros)

    cursor.execute(query, valores)
    resultados = cursor.fetchall()

    # ===============================
    # TRANSFORMACIÓN DE COORDENADAS
    # ===============================
    transformadores: dict[str, Transformer] = {}
    coordenadas = []

    for fila in resultados:
        try:
            lat = str(fila.get('Latitud_decimal', '')).replace(',', '.')
            lon = str(fila.get('Longitud_decimal', '')).replace(',', '.')
            epsg = fila.get('Codigo_EPSG_decimal')
            if lat and lon and epsg:
                lat = float(lat); lon = float(lon)
                if epsg not in transformadores:
                    transformadores[epsg] = Transformer.from_crs(
                        f"EPSG:{epsg}", "EPSG:4326", always_xy=True
                    )
                lon_wgs84, lat_wgs84 = transformadores[epsg].transform(lon, lat)
                # NO sobrescribir los valores decimales originales (proyectados)
                fila['Longitud_mapa'] = lon_wgs84
                fila['Latitud_mapa'] = lat_wgs84
                coordenadas.append({
                    'lat': lat_wgs84,
                    'lon': lon_wgs84,
                    'Nombre_cientifico': fila.get('Nombre_cientifico') or 'No disponible',
                    'Nombre_comun': fila.get('Nombre_comun') or 'No disponible',
                    'Codigo_de_muestra': fila.get('Codigo_de_muestra') or 'No disponible',
                    'Proyecto': fila.get('Proyecto') or 'No disponible',
                    'Fecha_de_colecta': fila.get('Fecha_de_colecta') or 'No disponible'
                })
        except:
            fila['Latitud_mapa'] = None
            fila['Longitud_mapa'] = None

    # =======================================
    # EXPORTACIÓN DE RESULTADOS A CSV Y EXCEL
    # =======================================
    # Columnas finales de exportación (alineadas entre CSV y Excel)
    include_map = os.getenv("EXPORT_INCLUDE_MAP_COORDS", "1") == "1"
    export_columnas = columnas_mostrar.copy()
    if include_map:
        # Solo agregamos si existen coordenadas calculadas y no están ya en la lista
        if any(f.get('Latitud_mapa') is not None for f in resultados):
            if 'Latitud_mapa' not in export_columnas:
                export_columnas.append('Latitud_mapa')
        if any(f.get('Longitud_mapa') is not None for f in resultados):
            if 'Longitud_mapa' not in export_columnas:
                export_columnas.append('Longitud_mapa')

    export_id = str(uuid.uuid4())
    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    csv_path = os.path.join(app.config['EXPORT_FOLDER'], f"{export_id}.csv")
    xlsx_path = os.path.join(app.config['EXPORT_FOLDER'], f"{export_id}.xlsx")

    # =====================
    # EXPORTACIÓN A CSV
    # =====================
    add_bom = os.getenv("CSV_ADD_BOM", "1") == "1"
    delimiter = os.getenv("CSV_DELIMITER", ",")
    quoting_mode = csv.QUOTE_MINIMAL
    fecha_cols = {c for c in export_columnas if 'fecha' in c.lower()}
    total_registros = len(resultados)
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        if add_bom:
            f.write('\ufeff')
        # Línea de metadatos inicial
        f.write(f"# total_registros: {total_registros}\n")
        writer = csv.DictWriter(
            f,
            fieldnames=export_columnas,
            delimiter=delimiter,
            quoting=quoting_mode
        )
        writer.writeheader()
        for fila in resultados:
            out_row = {}
            for col in export_columnas:
                val = fila.get(col, '')
                if col in fecha_cols and isinstance(val, str):
                    parsed = None
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
                        try:
                            parsed = datetime.strptime(val, fmt)
                            break
                        except Exception:
                            pass
                    if parsed:
                        val = parsed.strftime('%Y-%m-%d')
                out_row[col] = val
            writer.writerow(out_row)

    # =====================
    # EXPORTACIÓN A EXCEL
    # =====================
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "Datos"
    ws.append(export_columnas)

    for fila in resultados:
        row_values = []
        for col in export_columnas:
            val = fila.get(col, '')
            if col in fecha_cols and isinstance(val, str):
                parsed = None
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
                    try:
                        parsed = datetime.strptime(val, fmt)
                        break
                    except Exception:
                        pass
                if parsed:
                    val = parsed
            row_values.append(val)
        ws.append(row_values)

    header_fill_color = os.getenv('EXCEL_HEADER_FILL', '18263f')
    header_font_color = os.getenv('EXCEL_HEADER_FONT', 'e6ebff')
    for cell in ws[1]:
        cell.font = Font(bold=True, color=header_font_color)
        cell.fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    max_width = int(os.getenv('EXCEL_MAX_COL_WIDTH', '60'))
    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells:
            val = c.value
            if val is None:
                length = 0
            elif isinstance(val, datetime):
                length = len(val.strftime('%Y-%m-%d'))
            else:
                length = len(str(val))
            if length > max_len:
                max_len = length
        adjusted = min(max_len + 2, max_width)
        ws.column_dimensions[get_column_letter(idx)].width = adjusted
    for col in fecha_cols:
        try:
            idx = export_columnas.index(col) + 1
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=idx)
                if isinstance(cell.value, datetime):
                    cell.number_format = 'YYYY-MM-DD'
        except ValueError:
            pass
    # Hoja resumen
    resumen = wb.create_sheet(title="Resumen")
    resumen.append(["Campo", "Valor"])
    resumen.append(["total_registros", total_registros])
    resumen.append(["fecha_exportacion", timestamp_str])
    resumen.append(["columnas", ",".join(export_columnas)])
    resumen.append(["incluir_coord_mapa", "si" if include_map else "no"])
    # Estilo simple para encabezado resumen
    resumen["A1"].font = Font(bold=True)
    resumen["B1"].font = Font(bold=True)
    wb.save(xlsx_path)

    # Guardar rutas en sesión (Flask session SOLO para paths de exportación)
    session['csv_export_path'] = csv_path
    session['excel_export_path'] = xlsx_path
    session['export_columnas'] = export_columnas
    session['export_timestamp'] = timestamp_str

    cursor.close()
    conn.close()

    return render_template(
        'results.html',
        resultados=resultados,
        columnas_mostrar=columnas_mostrar,
        columnas_csv=columnas_mostrar,
        palabra=palabra_clave,
        columna=columna_clave,
        coordenadas=coordenadas,
        **brand_vars()
    )

# ===================================
# RUTA "/exportar_csv" (DESCARGA CSV)
# ===================================
@app.route('/exportar_csv')
def exportar_csv():
    export_path = session.get('csv_export_path')
    if not export_path or not os.path.exists(export_path):
        return 'No hay resultados para exportar en CSV', 400
    def generate():
        with open(export_path, 'r', encoding='utf-8') as f:
            for line in f:
                yield line
    ts = session.get('export_timestamp', datetime.now().strftime('%Y%m%d_%H%M%S'))
    resp = Response(generate(), mimetype='text/csv')
    resp.headers['Content-Disposition'] = f'attachment; filename=resultados_{ts}.csv'
    return resp

# =======================================
# RUTA "/exportar_excel" (DESCARGA EXCEL)
# =======================================
@app.route('/exportar_excel')
def exportar_excel():
    export_path = session.get('excel_export_path')
    if not export_path or not os.path.exists(export_path):
        return 'No hay resultados para exportar en Excel', 400
    ts = session.get('export_timestamp', datetime.now().strftime('%Y%m%d_%H%M%S'))
    return send_file(
        export_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'resultados_{ts}.xlsx'
    )

# (Opcional) endpoint de salud
@app.route('/health')
def health():
    try:
        conn = cnxpool.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        cursor.fetchone()
        cursor.close()
        conn.close()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}, 500

# ===============================
# EJECUCIÓN DE LA APLICACIÓN
# ===============================
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
